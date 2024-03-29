import requests
import os
from django.shortcuts import render, redirect
from main.defs import *
from main.models import *
from datetime import datetime, timedelta
from pytz import timezone, utc
from openpyxl import load_workbook   # for xls? formats
import xlrd  # for xls formats
from django.contrib import messages
import rdflib
from rdflib import Namespace, URIRef, Literal
from rdflib.namespace import RDF, RDFS, OWL, FOAF

from urllib.request import urlopen

def index(request):
    return render(request, 'index.html')

def importa_dados(request):
    sites = ProvenanceStatement.objects.all().order_by('id')
    context = {
        'sites': sites,
    }

    return render(request, 'importa_dados.html', context)

def raspagem_curso(request):

    os.system('scrapy runspider scrapies_esp/scrapy_curso.py')

    return redirect('/index')

def raspagem_job(request):

    os.system('scrapy runspider scrapies_esp/scrapy_job_net_empregos.py')

    return redirect('/index')

def visualiza_ontologia(request):
    return render(request, 'visualiza_ontologia.html')

def sparql_ontologia(request):

    context = {
        'sparql_query': '',
        'sparql_resultado': 'Sem resultados para mostrar.',
    }

    return render(request, 'sparql_ontologia.html', context)

def serializa_ontologia(request):
    try:
        g = processa_ontologia()

        ontologia_serializada = {}
        ontologia_serializada['n3'] = g.serialize(format='n3').decode('utf-8')
        ontologia_serializada['nt'] = g.serialize(format='nt').decode('utf-8')
        ontologia_serializada['prettyxml'] = g.serialize(format='pretty-xml').decode('utf-8')
        ontologia_serializada['trig'] = g.serialize(format='trig').decode('utf-8')
        ontologia_serializada['turtle'] = g.serialize(format='turtle').decode('utf-8')
        ontologia_serializada['xml'] = g.serialize(format='xml').decode('utf-8')

        context = {
            'ontologia_serializada': ontologia_serializada,
        }

        messages.success(request, 'Serialização da ontologia executada com sucesso.')

    except Exception as error:
        messages.error(request, 'Ocorreu um erro ao serializar a ontologia. Operação não efetuada.')

    return render(request, 'serializa_ontologia.html', context)

def run_sparql_ontologia(request):
    try:
        g = processa_ontologia()
        sparql_query = str(request.POST.get('sparql_text', '').strip())
        try:
            resultado = g.query(sparql_query)
            # resultado = g.query('SELECT ?name WHERE { ?label owl:NamedIndividual ?name }')
            # resultado = g.query('SELECT ?x ?label WHERE { ?x owl:NamedIndividual ?label }')
        except:
            resultado = None

        print('resultado', resultado)

        sparql_resultado = []
        if resultado:
            for x in resultado:
                print(x)
                sparql_resultado.append(str(x) + '\n')

        if not sparql_resultado:
            sparql_resultado = ['Ocorreu um erro ao executar a query. \n\nSem resultados para mostrar.']
            messages.error(request, 'Ocorreu um erro ao executar a query. Operação não efetuada.')
        else:
            messages.success(request, 'Query executada com sucesso.')

        context = {
            'sparql_query': sparql_query,
            'sparql_resultado': sparql_resultado,
        }

    except Exception as error:
        messages.error(request, 'Ocorreu um erro ao executar a query. Operação não efetuada.')

    return render(request, 'sparql_ontologia.html', context)

def processa_ontologia():

    myOntology = Namespace('https://github.com/LuSoMaBra/semantic-esp/tree/master/semanticesp/ontology#')

    # CURSO
    g = rdflib.Graph()
    ontologia = g.parse('ontology/ontologia_psesp_rdf.owl')
    g.bind('myOntology', myOntology)
    curso = Curso.objects.select_related(None).all()
    trabalho = Trabalho.objects.select_related(None).all()
    last_extraction = curso.distinct('provenance_statement__last_extraction').order_by('-provenance_statement__last_extraction').first().provenance_statement.last_extraction
    last_extraction_trabalho = trabalho.distinct('provenance_statement__last_extraction').order_by('-provenance_statement__last_extraction').first().provenance_statement.last_extraction
    if last_extraction:
        curso = curso.filter(provenance_statement__last_extraction__gte=last_extraction)

    for x in curso:
        # trabalho
        trabalho = trabalho.filter(provenance_statement__last_extraction__gte=last_extraction_trabalho, area_curso=x.curso_cnaef.areacnaef)
        for y in trabalho:
            g.add((myOntology['title'], OWL.NamedIndividual, Literal(y.titulo)))
            g.add((myOntology['description'], OWL.NamedIndividual, Literal(y.descricao)))
            g.add((myOntology['baseSalary'], OWL.NamedIndividual, Literal(y.remuneracao)))
            g.add((myOntology['employmentType'], OWL.NamedIndividual, Literal(y.modo)))
            g.add((myOntology['qualifications'], OWL.NamedIndividual, Literal(y.area_curso)))

        # curso
        g.add((myOntology['name'], OWL.NamedIndividual, Literal(x.curso_cnaef.nome)))
        g.add((myOntology['name'], OWL.NamedIndividual, Literal(x.curso_cnaef.college_or_university.nomedoestabelecimento)))
        g.add((myOntology['description'], OWL.NamedIndividual, Literal(x.descricao)))
        g.add((myOntology['educationalProgramMode'], OWL.NamedIndividual, Literal(x.modo)))
        g.add((myOntology['url'], OWL.NamedIndividual, Literal(x.url)))
        g.add((myOntology['termDuration'], OWL.NamedIndividual, Literal(x.duracao)))
        g.add((myOntology['educationalCredentialAwarded'], OWL.NamedIndividual, Literal(x.curso_cnaef.niveldeformacao)))
        g.add((myOntology['programmeArea'], OWL.NamedIndividual, Literal(x.curso_cnaef.areacnaef)))
        g.add((myOntology['internationalRegistrationFee'], OWL.NamedIndividual, Literal(x.valor_propina_internacional)))
        g.add((myOntology['nationalRegistrationFee'], OWL.NamedIndividual, Literal(x.valor_propina_nacional)))

        # provenance_statement
        g.add((myOntology['title'], OWL.NamedIndividual, Literal(x.provenance_statement.title)))
        g.add((myOntology['url'], OWL.NamedIndividual, Literal(x.provenance_statement.url)))
        g.add((myOntology['creator'], OWL.NamedIndividual, Literal(x.provenance_statement.creator)))
        g.add((myOntology['created'], OWL.NamedIndividual, Literal(x.provenance_statement.created)))
        g.add((myOntology['modified'], OWL.NamedIndividual, Literal(x.provenance_statement.modified)))
        g.add((myOntology['fileFormat'], OWL.NamedIndividual, Literal(x.provenance_statement.source)))
        g.add((myOntology['lastExtraction'], OWL.NamedIndividual, Literal(x.provenance_statement.last_extraction)))

        # college_or_university
        g.add((myOntology['branchCode'], OWL.NamedIndividual, Literal(x.curso_cnaef.college_or_university.codigodoestabelecimento)))
        g.add((myOntology['streetAddress'], OWL.NamedIndividual, Literal(x.curso_cnaef.college_or_university.morada)))
        g.add((myOntology['addressLocality'], OWL.NamedIndividual, Literal(x.curso_cnaef.college_or_university.concelho)))
        g.add((myOntology['addressRegion'], OWL.NamedIndividual, Literal(x.curso_cnaef.college_or_university.distrito)))
        g.add((myOntology['postalCode'], OWL.NamedIndividual, Literal(x.curso_cnaef.college_or_university.codigopostal)))

    # cursos = list(g.triples((myOntology['name'], OWL.NamedIndividual, None)))
    # for (sub, pred, obj) in cursos:
    #     print((sub, pred, obj))

    return g

def populate_child(request, id):

    # from scrapies_esp import db_tools
    # areas = db_tools.selectDB(db_tools.connectDB(), "select id, nome from curso_cnaef where estabelecimento like '{}__' and nome like '____ - {}' and niveldeformacao like 'Licenciatura%' ".format('12', 'Turismo'))
    #
    #
    # # curso_cnaef_id = db_tools.selectDB(db_tools.connectDB(),
    # #                                    # "select now()")
    # #                                    "update provenance_statement set last_extraction = now() where id = {};".format(9) + " select * from provenance_statement where id = {}; COMMIT;".format(9))
    # #
    #
    # # for x in curso_cnaef_id:
    # for x in areas:
    #     print(x[0])
    #
    # return redirect('/importa_dados')
    #

    try:
        provenance_statement = ProvenanceStatement.objects.get(id=id)
    except:
        provenance_statement = None
    if provenance_statement:
        populated = False
        try:
            # if provenance_statement.source == 'pdf':
            #     res = urlopen(provenance_statement.url)
            # else:
            res = requests.get(provenance_statement.url)
        except Exception as error:
            res = None
            messages.error(request, 'Ocorreu um erro ao efetuar a requisiçao do arquivo {}. Operação não efetuada.'.format(provenance_statement.codigo))
            messages.info(request, 'Erro: {}'.format(error))
        try:
            if (res) and (provenance_statement.source == 'json'):
                if provenance_statement.codigo == 'InstituicoesdoEnsinoSuperior':
                    for defaults in (res.json()['d']):
                        partition_key = defaults['PartitionKey']
                        row_key = defaults['RowKey']
                        timestamp = (defaults['Timestamp'].split('T')[0] + ' ' + defaults['Timestamp'].split('T')[1].split('.')[0])
                        modified = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        defaults.pop('PartitionKey')
                        defaults.pop('RowKey')
                        defaults.pop('Timestamp')
                        codigodoestabelecimento = '{:0>4}'.format(defaults['codigodoestabelecimento'])
                        defaults.pop('codigodoestabelecimento')
                        populated, c = CollegeOrUniversity.objects.update_or_create(
                            partitionkey = partition_key,
                            rowkey = row_key,
                            provenance_statement = provenance_statement,
                            timestamp = timestamp,
                            codigodoestabelecimento = codigodoestabelecimento,
                            defaults = defaults
                        )

                        # DEPENDE DA RASPAGEM
                        # IMPLEMENTAR CAMPO LINKED_DATA_UNIVERSIDADE

                elif provenance_statement.codigo == 'ClassNacionaldeareasdeeducacaoeformacao':
                    for defaults in (res.json()['d']):
                        partition_key = defaults['PartitionKey']
                        row_key = defaults['RowKey']
                        timestamp = (defaults['Timestamp'].split('T')[0] + ' ' + defaults['Timestamp'].split('T')[1].split('.')[0])
                        modified = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        defaults.pop('PartitionKey')
                        defaults.pop('RowKey')
                        defaults.pop('Timestamp')
                        defaults['estabelecimento'] = defaults['estabelecimento'].split(' - ')[0]
                        college_or_university = CollegeOrUniversity.objects.get(codigodoestabelecimento=defaults['estabelecimento'])
                        defaults['nome'] = defaults['curso']
                        defaults.pop('curso')
                        populated, c = CursoCnaef.objects.update_or_create(
                            partitionkey = partition_key,
                            rowkey = row_key,
                            provenance_statement = provenance_statement,
                            timestamp = timestamp,
                            college_or_university = college_or_university,
                            defaults = defaults
                        )
                else:
                    pass

            elif (res) and (provenance_statement.source == 'html'):
                if provenance_statement.codigo == 'UTAD_spider':
                    os.system('scrapy runspider scrapies_esp/scrapy_curso.py')
                    modified = provenance_statement.created
                    populated = True

                if provenance_statement.codigo == 'net-empregos':
                    os.system('scrapy runspider scrapies_esp/scrapy_job_net_empregos.py')
                    modified = provenance_statement.created
                    populated = True

            elif (res) and (provenance_statement.source == 'xls'):

                file = open('tmp_excel_work.xls', 'wb')
                file.write(res.content)
                file.close()

                book = xlrd.open_workbook(file.name)
                print("The number of worksheets is {0}".format(book.nsheets))
                print("Worksheet name(s): {0}".format(book.sheet_names()))
                sheet = book.sheet_by_index(0)
                print("{0} {1} {2}".format(sheet.name, sheet.nrows, sheet.ncols))
                print("Cell A2 is {0}".format(sheet.cell_value(rowx=1, colx=0)))
                for rx in range(sheet.nrows):
                    print(sheet.cell_value(rx,0), sheet.cell_value(rx,1))
                    # print(sheet.row(rx)[0])
                os.remove(file.name)

            elif (res) and (provenance_statement.source == 'xls?'):
                file = open('tmp_excel_work.xlsx', 'wb')
                file.write(res.content)

                wb = load_workbook(file, False, False, False, False)
                sheet = wb.worksheets[0]  # aba da planilha

                max_col = sheet.max_column + 1
                max_row = sheet.max_row + 1

                for lin in range(13, max_col):
                    celula1 = sheet.cell(row=lin, column=1)
                    celula2 = sheet.cell(row=lin, column=2)
                    print(celula1,celula2)
                wb.close()
                os.remove(file.name)

            elif (res) and (provenance_statement.source == 'pdf'):

                file = open('tmp_pdf_work.pdf', 'wb')
                file.write(res.content)
                file.close()
                # stringSaida = lerPDF(file)
                stringSaida = getPDFContent(file.name)
                txt = stringSaida.split('\n')
                print(txt[222])
                # for x in stringSaida.split('\n'):
                #     print(x)
                # os.remove(file.name)
            else:
                pass
            if populated:   # deve atualizar essa variável para atualização do provenance_statement
                provenance_statement.creator = 'Portal de dados abertos da Administração Pública (https://dados.gov.pt)'
                provenance_statement.created = modified
                provenance_statement.modified = modified
                provenance_statement.last_extraction = datetime.now(tz=utc)
                provenance_statement.populated = True
                provenance_statement.save()
            messages.success(request, 'Processamento do Arquivo: "{}" efetuado com sucesso.'.format(provenance_statement.codigo))
        except Exception as error:
            messages.error(request, 'Ocorreu um erro ao processar o arquivo {}. Operação não efetuada.'.format(provenance_statement.codigo))
            messages.info(request, 'Erro: {}'.format(error))

    return redirect('/importa_dados')



